"""朝イチ候補のアフター検証 — GitHub Actions版（保険用）

run_research.py が朝に生成した reports/YYYY-MM-DD.md の朝イチ候補を、
当日の実際の値動きで答え合わせし、reports/verification-ledger.csv に
追記、reports/朝イチ検証台帳.xlsx を再生成する。

必要な環境変数:
  ANTHROPIC_API_KEY  console.anthropic.com で発行したAPIキー
  RESEARCH_MODEL     省略時 claude-opus-4-8
"""

import csv
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import anthropic

JST = timezone(timedelta(hours=9))
REPO_ROOT = Path(__file__).resolve().parent.parent
REPORTS_DIR = REPO_ROOT / "reports"
LEDGER_CSV = REPORTS_DIR / "verification-ledger.csv"
LEDGER_XLSX = REPORTS_DIR / "朝イチ検証台帳.xlsx"

MODEL = os.environ.get("RESEARCH_MODEL", "claude-opus-4-8")
MAX_CONTINUATIONS = 8

COLUMNS = [
    "日付", "銘柄", "コード", "材料", "推奨区分",
    "前日終値", "始値", "高値", "安値", "終値", "GAP率%",
    "A:寄り成り損益%", "A:逆指値発動", "B:教科書ルール損益%(近似)",
    "寄り天?", "教科書ルール判定", "メモ",
]
JSON_KEYS = [
    "date", "name", "code", "catalyst", "category",
    "prev_close", "open", "high", "low", "close", "gap_pct",
    "a_pnl_pct", "a_stopped", "b_pnl_pct",
    "yorten", "verdict", "memo",
]

PROMPT = """あなたは日本株のアフター検証アシスタントです。今日は日本時間 {today} です。
下に貼る「今朝のモーニングリサーチ」の朝イチ候補リスト（見送り推奨も含む）の各銘柄について、
Web検索で本日の四本値（前日終値・始値・高値・安値・終値）を調べ、次の3シナリオを計算してください。

A) 寄り成り買い+逆指値-3%: 始値で買い、安値≦始値×0.97なら-3%で損切り（高安の順序不明なら保守的に発動扱い）。未発動なら(終値-始値)/始値。
B) 教科書ルール（寄り後15分様子見→押し目指値）の近似: (始値+安値)÷2で買えた仮定、引け売り、-3%逆指値適用。近似である旨をメモに書く。
C) 見送り推奨銘柄: 買っていたらの損益をAで計算し、見送りが正解だったか判定。

データが取れない銘柄は数値を null にして memo に「取得不可」と書く。推測で数値を埋めないこと。
東証が休場の日は空配列 [] だけを出力すること。

最終出力は JSON 配列のみ（コードフェンスなし・前後の文章なし）。各要素のキー:
{keys}
数値はパーセントも含めて数値型で（例: gap_pct は +10.0% なら 10.0）。a_stopped は "発動"/"未発動"、
category は "候補"/"見送り推奨"、yorten は "寄り天"/"寄り天でない"/"不明"。

----- 今朝のレポート -----
{report}
"""


def load_morning_report() -> str | None:
    path = REPORTS_DIR / f"{datetime.now(JST).strftime('%Y-%m-%d')}.md"
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8")


def ask_claude(report: str) -> list[dict]:
    client = anthropic.Anthropic()
    tools = [{"type": "web_search_20260209", "name": "web_search", "max_uses": 25}]
    prompt = PROMPT.format(
        today=datetime.now(JST).strftime("%Y年%m月%d日"),
        keys=", ".join(JSON_KEYS),
        report=report,
    )
    messages = [{"role": "user", "content": prompt}]

    response = None
    for _ in range(MAX_CONTINUATIONS):
        with client.messages.stream(
            model=MODEL,
            max_tokens=32000,
            thinking={"type": "adaptive"},
            tools=tools,
            messages=messages,
        ) as stream:
            response = stream.get_final_message()
        if response.stop_reason != "pause_turn":
            break
        messages = [
            {"role": "user", "content": prompt},
            {"role": "assistant", "content": response.content},
        ]

    if response is None or response.stop_reason == "refusal":
        raise RuntimeError(f"検証の応答を取得できませんでした: {getattr(response, 'stop_reason', None)}")

    text = "\n".join(b.text for b in response.content if b.type == "text").strip()
    # コードフェンスや前置きが混ざった場合に備えてJSON配列部分を抽出
    m = re.search(r"\[.*\]", text, flags=re.S)
    if not m:
        raise RuntimeError(f"JSON配列が見つかりません: {text[:300]}")
    return json.loads(m.group(0))


def append_to_csv(rows: list[dict]) -> None:
    REPORTS_DIR.mkdir(exist_ok=True)
    is_new = not LEDGER_CSV.exists()
    with LEDGER_CSV.open("a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if is_new:
            writer.writerow(COLUMNS)
        for r in rows:
            writer.writerow([r.get(k, "") for k in JSON_KEYS])


def rebuild_xlsx() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    with LEDGER_CSV.open(newline="", encoding="utf-8-sig") as f:
        records = list(csv.reader(f))

    wb = Workbook()
    ws = wb.active
    ws.title = "検証履歴"
    for row in records:
        ws.append(row)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F4E79")
    ws.freeze_panes = "A2"

    # サマリー（データ行から集計。数値変換できない値は無視）
    def nums(col_idx):
        out = []
        for row in records[1:]:
            try:
                out.append(float(row[col_idx]))
            except (ValueError, IndexError):
                pass
        return out

    a_pnl, b_pnl = nums(11), nums(13)
    candidates = [r for r in records[1:] if len(r) > 4 and r[4] == "候補"]
    skips = [r for r in records[1:] if len(r) > 4 and r[4] == "見送り推奨"]
    skip_correct = sum(1 for r in skips if len(r) > 15 and "正解" in r[15])

    ws2 = wb.create_sheet("サマリー")
    days = len({r[0] for r in records[1:] if r})
    summary = [
        ["朝イチ候補 検証サマリー", ""],
        ["累計検証日数", days],
        ["累計検証銘柄数", len(records) - 1],
        ["候補銘柄数 / 見送り推奨数", f"{len(candidates)} / {len(skips)}"],
        ["A(寄り成り) 勝率", f"{sum(1 for x in a_pnl if x > 0)}/{len(a_pnl)}" if a_pnl else "-"],
        ["A 平均損益%", round(sum(a_pnl) / len(a_pnl), 2) if a_pnl else "-"],
        ["B(教科書ルール近似) 勝率", f"{sum(1 for x in b_pnl if x > 0)}/{len(b_pnl)}" if b_pnl else "-"],
        ["B 平均損益%", round(sum(b_pnl) / len(b_pnl), 2) if b_pnl else "-"],
        ["教科書ルールの優位性(B-A)%",
         round(sum(b_pnl) / len(b_pnl) - sum(a_pnl) / len(a_pnl), 2) if a_pnl and b_pnl else "-"],
        ["見送り推奨の的中", f"{skip_correct}/{len(skips)}" if skips else "-"],
    ]
    for r in summary:
        ws2.append(r)
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.column_dimensions["A"].width = 34

    ws3 = wb.create_sheet("前提条件")
    for n in [
        "シナリオA: 始値買い・安値が-3%到達で逆指値発動(-3%固定計上。実際はスリッページでより悪化しうる)",
        "シナリオB: (始値+安値)÷2で約定した近似・引け売り・-3%逆指値。分足未使用の近似値",
        "手数料・税・スリッページ未考慮。株価はWeb検索由来のため誤りうる",
        "本検証はシミュレーションであり実際の約定を保証しません。投資判断はご自身の責任で行ってください",
    ]:
        ws3.append([n])
    ws3.column_dimensions["A"].width = 95

    wb.save(LEDGER_XLSX)


def main() -> None:
    report = load_morning_report()
    if report is None:
        print("今朝のレポートが reports/ に無いためスキップ（朝のワークフロー未実行か休場）")
        return
    rows = ask_claude(report)
    if not rows:
        print("休場または候補なしのため記録なし")
        return
    append_to_csv(rows)
    rebuild_xlsx()
    print(f"{len(rows)}銘柄を検証し台帳を更新: {LEDGER_XLSX}", file=sys.stderr)


if __name__ == "__main__":
    main()
